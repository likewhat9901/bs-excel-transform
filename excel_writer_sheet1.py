from openpyxl import load_workbook


def load_excel_workbook(file_path):
    """엑셀 파일 열기"""
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        print(f"시트 이름: {ws.title}")
        print(f"시트 이름 목록: {wb.sheetnames}")
        return wb, ws
    except Exception as e:
        print(f"❌ 파일 열기 실패: {e}")
        return None, None


def validate_ranges(sorted_income, sorted_expense, 
                   income_start_row, income_end_row, 
                   expense_start_row, expense_end_row):
    """데이터 범위 검증"""
    income_range = income_end_row - income_start_row
    expense_range = expense_end_row - expense_start_row
    
    if (len(sorted_income) - 1) != income_range:
        print(f"⚠️ 경고: 수입 데이터({len(sorted_income)-1}개)가 범위({income_range}개)와 일치하지 않습니다!")
        return False
    
    if (len(sorted_expense) - 1) != expense_range:
        print(f"⚠️ 경고: 지출 데이터({len(sorted_expense)-1}개)가 범위({expense_range}개)와 일치하지 않습니다!")
        return False
    
    print(f"✓ 검증 완료: 수입({income_range}행), 지출({expense_range}행)")
    return True


def write_data_to_sheet(ws, data, start_row):
    """DataFrame 데이터를 엑셀 시트에 저장"""
    row_start = start_row + 1  # Pandas Index 0 -> Excel 행 1
    for i in range(len(data)):
        row_num = row_start + i
        for col_num, value in enumerate(data.iloc[i].values, start=1):
            ws.cell(row=row_num, column=col_num, value=value)


def save_to_excel(file_path, sorted_income, sorted_expense, 
                  income_start_row, income_end_row, 
                  expense_start_row, expense_end_row):
    """정렬된 데이터를 엑셀 파일에 저장"""
    # 1. 파일 열기
    wb, ws = load_excel_workbook(file_path)
    if wb is None:
        return False
    
    # 2. 범위 검증
    if not validate_ranges(sorted_income, sorted_expense,
                          income_start_row, income_end_row,
                          expense_start_row, expense_end_row):
        return False
    
    # 3. 데이터 저장
    try:
        write_data_to_sheet(ws, sorted_income, income_start_row)
        write_data_to_sheet(ws, sorted_expense, expense_start_row)
        
        output_file = file_path
        wb.save(output_file)
        print(f"✓ 파일 저장 완료: {output_file}")
        return True
    except Exception as e:
        print(f"❌ 파일 저장 중 오류 발생: {e}")
        return False
