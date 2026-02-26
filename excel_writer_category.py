import pandas as pd
from openpyxl.styles import PatternFill, Alignment, Font, Border

# 시트별 합계 계산
def calculate_total_amount(df):
    """금액 합계 계산"""
    total = pd.to_numeric(df['금액'].astype(str).str.replace(',', ''), errors='coerce').sum()
    return total

# 합계 행 추가
def add_summary_row(df, total_amount):
    """합계 행을 DataFrame 하단에 추가"""
    summary_row = {col: "" for col in df.columns}
    summary_row['날짜'] = "월 총계"
    summary_row['금액'] = total_amount
    summary_row['화폐'] = "KRW"
    
    df = pd.concat([df, pd.DataFrame([summary_row])], ignore_index=True)
    return df

# 데이터 포맷 변경
def format_data(df):
    """날짜, 시간 포맷 변환"""
    df = df.copy()
    
    if '날짜' in df.columns:
        df['날짜'] = df['날짜'].dt.strftime('%Y-%m-%d')
    if '시간' in df.columns:
        df['시간'] = pd.to_datetime(df['시간'], format='%H:%M:%S').dt.strftime('%H:%M')
    return df

# 헤더 스타일 적용
def apply_header_style(ws, df):
    """헤더 배경색, 정렬, 너비 설정"""
    header_fill = PatternFill(start_color='7B808D', end_color='7B808D', fill_type='solid')
    header_font = Font(color="000000", bold=False)
    center_alignment = Alignment(horizontal='center', vertical='center')
    no_border = Border()

    column_widths = {
        "날짜": 16, "시간": 9, "타입": 9,
        "대분류": 9, "소분류": 9, "내용": 36,
        "금액": 21, "화폐": 9, "결제수단": 36, "메모": 36
    }

    for i, col in enumerate(df.columns, start=1):
        col_name = str(col)                                         # 컬럼명 문자열 변환
        cell = ws.cell(row=1, column=i)                             # 헤더 행의 i번째 열 선택

        cell.fill = header_fill                                     # 헤더 배경색 적용
        cell.font = header_font                                     # 헤더 글꼴 적용
        cell.alignment = center_alignment                           # 헤더 정렬 적용
        cell.border = no_border                                     # 테두리 없음

        width = column_widths.get(col_name, 15)                     # 컬럼 너비 가져오기
        ws.column_dimensions[cell.column_letter].width = width      # 컬럼 너비 설정

# 데이터 정렬 적용
def apply_data_alignment(ws, df):
    """모든 데이터 행에 대해 정렬 및 숫자 포맷 적용"""
    column_alignments = {
        # 기본값 = left
        "날짜": Alignment(horizontal='right'),
        "시간": Alignment(horizontal='right'),
        "금액": Alignment(horizontal='right'),
    }

    for i, col in enumerate(df.columns, start=1):
        col_name = str(col)
        alignment = column_alignments.get(col_name, Alignment(horizontal='left'))
        
        # 2행부터 마지막 데이터 행까지 반복
        for row_num in range(2, len(df) + 2):
            cell = ws.cell(row=row_num, column=i)
            cell.alignment = alignment

            # 금액 컬럼인 경우 천 단위 콤마 서식 적용
            if col_name == "금액":
                cell.number_format = '#,##0'

# 합계 행 스타일 적용
def apply_summary_style(ws, df):
    """합계 행(맨 마지막)에 굵은 글씨 적용"""
    last_row = len(df) + 1  # 엑셀 행 번호 (1부터 시작)
    for col_num in range(1, len(df.columns) + 1):
        ws.cell(row=last_row, column=col_num).font = Font(bold=True)

# 시트 생성 (통합)
def create_sheet(writer, sheet_name, df):
    """하나의 시트 생성 및 스타일 적용"""
    # --- [STEP 1: 데이터 가공] ---
    df_copy = df.copy()
    
    # 1. 합계 계산 및 행 추가 (엑셀에 쓰기 전 모든 데이터 준비 완료)
    total_amount = calculate_total_amount(df_copy)
    
    # 2. 날짜/시간 포맷팅
    df_formatted = format_data(df_copy)

    # 3. 합계 행 추가
    df_final = add_summary_row(df_formatted, total_amount)
    
    # --- [STEP 2: 엑셀 파일 쓰기] ---
    df_final.to_excel(writer, sheet_name=sheet_name, index=False, header=True)
    ws = writer.sheets[sheet_name]
    
    # --- [STEP 3: 스타일 및 서식 적용] ---
    # 1. 헤더 스타일 (배경색, 너비 등)
    apply_header_style(ws, df_final)
    
    # 2. 데이터 정렬 및 숫자 포맷 (합계 행 포함 전체 적용)
    apply_data_alignment(ws, df_final)
    
    # 3. 합계 행 강조 (맨 마지막 줄만 별도 처리)
    apply_summary_style(ws, df_final)
    
    # 실제 데이터 건수는 합계 행을 제외한 개수로 출력
    print(f"✓ 시트 생성 완료: {sheet_name} (내역 {len(df)}건 + 합계 1건)")

# 메인 함수
def add_classified_sheets(file_path, classified_data):
    """분류된 데이터를 엑셀 시트로 추가"""
    try:
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            print(f"\n=== 시트 추가 시작 ===")
            for sheet_name, df in classified_data.items():
                create_sheet(writer, sheet_name, df)
        
        print(f"\n✨ 모든 분류 데이터가 '{file_path}'에 추가되었습니다.")
        return True
    except Exception as e:
        print(f"❌ 시트 추가 중 오류 발생: {e}")
        return False